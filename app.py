from flask import Flask, render_template, request, session, redirect, url_for,flash, Response
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from datetime import date,datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

app = Flask(__name__)

#Config
app.config['SECRET_KEY'] = os.getenv("SESSION_KEY")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

#DB Layout
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    assigned_tasks = db.relationship('Task',foreign_keys='Task.assigned_to',backref='assignee',lazy=True) # Relationships
    created_tasks = db.relationship('Task',foreign_keys='Task.assigned_by',backref='creator',lazy=True) # Relationships

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(850),nullable=False)
    assigned_to = db.Column(db.Integer,db.ForeignKey('user.id'),nullable=False)
    assigned_by = db.Column(db.Integer,db.ForeignKey('user.id'),nullable=False)
    project_id = db.Column(db.Integer,db.ForeignKey('project.id'),nullable=True)
    status = db.Column(db.String(20),default="Pending")
    created_date = db.Column(db.Date,default=date.today)
    due_date = db.Column(db.Date,nullable=False)
    completed_date = db.Column(db.Date,nullable=True)

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200),nullable=False)
    assigned_to = db.Column(db.Integer,db.ForeignKey('user.id'),nullable=False)
    assigned_by = db.Column(db.Integer,db.ForeignKey('user.id'),nullable=False)
    # created_date = db.Column(db.Date,default=date.today)
    tasks = db.relationship('Task',backref='project',lazy=True,cascade='all, delete')
    assignee = db.relationship('User',foreign_keys=[assigned_to],backref='assigned_projects')
    creator = db.relationship('User',foreign_keys=[assigned_by],backref='created_projects')

# with app.app_context():
#     db.create_all() #--> this increases cold start latency by 2-5 secs , thus create DB Manually

#Login
@app.route('/', methods=['GET','POST'])
def login():
    if "user_id" in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        try:
            user = User.query.filter_by(email=email).first()

            if user and bcrypt.check_password_hash(user.password,password):
                session["user_id"] = user.id
                return redirect(url_for('dashboard'))
            else:
                flash("Invalid Credentials. Please try again.")
                return render_template('login.html')
        except:
            flash('Internal Application Error')
            return render_template('login.html')
    return render_template('login.html')

# Registration
@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        try:
            userName = request.form.get('full_name')
            email = request.form.get('email')
            password = bcrypt.generate_password_hash(request.form.get('password')).decode('utf-8')

            new_user = User(name=userName,email=email,password=password)
        
            db.session.add(new_user)
            db.session.commit()
            return redirect(url_for('login'))
        except:
            flash('Please Check information or try again later.')
            return render_template('register.html')
    return render_template('register.html')

#Export Data
@app.route('/export-data')
def export_data():
    if "user_id" not in session:
        return redirect(url_for("login"))
 
    user_id = session["user_id"]
    wb = openpyxl.Workbook()
 
    # --- Sheet 1: Projects ---
    ws_p = wb.active
    ws_p.title = "Projects"
    ws_p.append(["Project Name", "Assigned By", "Assigned To", "Total Tasks", "Pending", "Completed", "Created Date", "Due Date", "Completed Date"])
    for cell in ws_p[1]:
        cell.font = Font(bold=True)
 
    # Styles for project summary rows vs nested task rows
    summary_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    summary_font = Font(bold=True)
    task_font = Font(italic=True, color="555555")
    indent_align = Alignment(indent=2)
 
    # Projects where I'm involved either way (assigned to me OR created by me)
    my_projects = Project.query.filter(
        (Project.assigned_to == user_id) | (Project.assigned_by == user_id)
    ).all()
 
    for p in my_projects:
        pending = sum(1 for t in p.tasks if t.status == "Pending")
        completed = sum(1 for t in p.tasks if t.status == "Completed")
 
        # --- Project summary row (bold, shaded) ---
        ws_p.append([
            p.name,
            p.creator.name,    # Assigned By
            p.assignee.name,   # Assigned To
            len(p.tasks),
            pending,
            completed,
            "", "", ""
        ])
        summary_row = ws_p.max_row
        for cell in ws_p[summary_row]:
            cell.font = summary_font
            cell.fill = summary_fill
 
        # --- Nested task rows (indented, italic) ---
        for t in p.tasks:
            ws_p.append([
                f"{t.title}",
                "", "",
                "",                                      # Total Tasks (blank for task rows)
                "Pending" if t.status == "Pending" else "",      # Pending column
                "Completed" if t.status == "Completed" else "",  # Completed column
                t.created_date.strftime("%Y-%m-%d") if t.created_date else "",
                t.due_date.strftime("%Y-%m-%d") if t.due_date else "",
                t.completed_date.strftime("%Y-%m-%d") if t.completed_date else ""
            ])
            task_row = ws_p.max_row
            ws_p.cell(row=task_row, column=1).font = task_font
            ws_p.cell(row=task_row, column=1).alignment = indent_align
            for col in range(2, 10):
                ws_p.cell(row=task_row, column=col).font = task_font
 
    # Widen first column so indentation + task titles are readable
    ws_p.column_dimensions['A'].width = 35
 
    # --- Sheet 2: Tasks ---
    ws_t = wb.create_sheet("Tasks")
    ws_t.append(["Title", "Project", "Assigned By", "Assigned To", "Status", "Created", "Due", "Completed"])
    for cell in ws_t[1]:
        cell.font = Font(bold=True)
 
    my_tasks = Task.query.filter(
        (Task.assigned_to == user_id) | (Task.assigned_by == user_id)
    ).all()
 
    for t in my_tasks:
        ws_t.append([
            t.title,
            t.project.name if t.project else "-",
            t.creator.name,    # Assigned By
            t.assignee.name,   # Assigned To
            t.status,
            t.created_date.strftime("%Y-%m-%d") if t.created_date else "",
            t.due_date.strftime("%Y-%m-%d") if t.due_date else "",
            t.completed_date.strftime("%Y-%m-%d") if t.completed_date else ""
        ])
 
    output = BytesIO()
    wb.save(output)
    output.seek(0)
 
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=my_tasks_export.xlsx"}
    )
 

#Dashboard
@app.route("/dashboard")
def dashboard():

    if "user_id" not in session:
        return redirect(url_for("login"))

    try:
        currentUser = User.query.get(session["user_id"])

        pending_tasks = Task.query.filter_by(
            assigned_to=currentUser.id,
            status="Pending",
            project_id=None
        ).all()

        completed_tasks = Task.query.filter_by(
            assigned_to=currentUser.id,
            status="Completed",
            project_id=None
        ).all()

        created_pending = Task.query.filter_by(
            assigned_by=currentUser.id,
            status="Pending",
            project_id=None
        ).all()

        created_completed = Task.query.filter_by(
            assigned_by=currentUser.id,
            status="Completed",
            project_id=None
        ).all()

        users = User.query.all()

        return render_template(
            "dashboard.html",
            currentUser=currentUser,
            users=users,
            pending=pending_tasks,
            completed=completed_tasks,
            created_pending=created_pending,
            created_completed=created_completed
        )
    except:
        flash('Error loading Data')
        return render_template(
        'dashboard.html',
        currentUser=None,
        users=[],
        pending=[],
        completed=[],
        created_pending=[],
        created_completed=[]
    )


@app.route('/add_task', methods=['POST'])
def add_task():

    if "user_id" not in session:
        return redirect(url_for("login"))

    title = request.form.get('title')
    assigned_to_name = request.form.get('to')

    assigned_user = User.query.filter_by(
        name=assigned_to_name
    ).first()

    try:
        due_date = datetime.strptime(
            request.form.get("due_date"),
            "%Y-%m-%d"
        ).date()
    except:
        flash('Please enter Target Date')
        return redirect(url_for('dashboard'))

    if title == None or title == "":
        flash('Please enter Task name')
        return redirect(url_for('dashboard'))

    if not assigned_user:
        flash("Please select user")
        return redirect(url_for('dashboard'))

    new_task = Task(
        title=title,
        assigned_to=assigned_user.id,
        assigned_by=session["user_id"],
        due_date=due_date
    )

    db.session.add(new_task)
    db.session.commit()

    flash("Task assigned successfully")

    return redirect(url_for('dashboard'))

@app.route('/complete-task/<int:task_id>', methods=['POST'])
def complete_task(task_id):

    if "user_id" not in session:
        return redirect(url_for("login"))

    task = Task.query.get_or_404(task_id)

    if task.assigned_to != session["user_id"]:
        flash("Unauthorized")
        return redirect(url_for("dashboard"))

    task.status = "Completed"
    task.completed_date = date.today()

    db.session.commit()

    return redirect(url_for('dashboard'))

@app.route('/complete-task-project/<int:task_id>', methods=['POST'])
def complete_task_project(task_id):
    if "user_id" not in session:
        return redirect(url_for("login"))

    task = Task.query.get_or_404(task_id)

    if task.assigned_to != session["user_id"]:
        flash("Unauthorized")
        return redirect(url_for("projects"))

    task.status = "Completed"
    task.completed_date = date.today()

    db.session.commit()

    return redirect(url_for('projects'))

# @app.route('/delete-task/<int:task_id>', methods=['POST'])
# def delete_task(task_id):

#     if "user_id" not in session:
#         return redirect(url_for("login"))

#     task = Task.query.get_or_404(task_id)

#     if task.assigned_to != session["user_id"]:
#         flash("Unauthorized action.")
#         return redirect(url_for("dashboard"))

#     db.session.delete(task)
#     db.session.commit()

#     flash("Task deleted successfully.")
#     return redirect(url_for("dashboard"))


@app.route('/projects', methods=['GET','POST'])
def projects():
    if "user_id" not in session:
        return redirect(url_for("login"))

    try:
        currentUser = User.query.get(session["user_id"])

        users = User.query.all()
        # active_projects = Project.query.filter_by(assigned_to=session["user_id"]).all()
        my_active_projects = []
        my_completed_projects = []
        allocated_active_projects = []
        allocated_completed_projects = []


        for project in Project.query.filter_by(assigned_to=session['user_id']).all():

            if any(task.status == "Pending" for task in project.tasks):
                my_active_projects.append(project)

            if any(task.status == "Completed" for task in project.tasks):
                my_completed_projects.append(project)
        
        for project in Project.query.filter_by(assigned_by=session['user_id']).all():
            if any(task.status == 'Pending' for task in project.tasks):
                allocated_active_projects.append(project)

            if any(task.status == 'Completed' for task in project.tasks):
                allocated_completed_projects.append(project)
        
        return render_template(
            "projects.html",
            currentUser=currentUser,
            users=users,
            my_active_projects=my_active_projects,
            my_completed_projects=my_completed_projects,
            allocated_active_projects=allocated_active_projects,
            allocated_completed_projects=allocated_completed_projects
        )
    except:
        flash('Error loading Data')
        return render_template(
            "projects.html",
            currentUser=None,
            users=[],
            my_active_projects=[],
            my_completed_projects=[],
            allocated_active_projects=[],
            allocated_completed_projects=[]
        )

@app.route('/add-project', methods=['GET','POST'])
def add_project():

    if "user_id" not in session:
        return redirect(url_for("login"))

    project_name = request.form.get("project_name")
    assigned_to_name = request.form.get("to")

    if not project_name:
        flash("Please enter project name")
        return redirect(url_for("projects"))

    assigned_user = User.query.filter_by(
        name=assigned_to_name
    ).first()

    if not assigned_user:
        flash("Please select a valid user")
        return redirect(url_for("projects"))

    # Get all tasks from modal
    task_titles = request.form.getlist("task_title[]")
    task_due_dates = request.form.getlist("task_due_date[]")

    valid_tasks = [t.strip() for t in task_titles if t.strip()]

    if not valid_tasks:
        flash("Please add at least one task")
        return redirect(url_for("projects"))

    try:
        project = Project(
            name=project_name,
            assigned_to=assigned_user.id,
            assigned_by=session["user_id"]
        )

        db.session.add(project)

        # Generate project.id
        db.session.flush()

        for title, due_date_str in zip(task_titles, task_due_dates):

            if not title.strip():
                continue

            due_date = datetime.strptime(
                due_date_str,
                "%Y-%m-%d"
            ).date()

            task = Task(
                title=title,
                assigned_to=assigned_user.id,
                assigned_by=session["user_id"],
                project_id=project.id,
                due_date=due_date
            )

            db.session.add(task)

        db.session.commit()

        flash("Project created successfully")

    except Exception as e:
        db.session.rollback()
        print(e)
        flash("Error creating project")

    return redirect(url_for("projects"))

#Logout
@app.route("/logout")
def logout():
    if "user_id" in session:
        session.clear()
        return redirect(url_for('login'))
    else:
        return redirect(url_for('login'))